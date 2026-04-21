"""Génère le fichier de suivi commercial des fins de bail triennales.

Optimisé mémoire pour Render Free (512 Mo) :
- Extraction en read_only, puis fermeture immédiate
- Construction de l'output dans un nouveau workbook léger
"""
from datetime import date, datetime
from io import BytesIO

import openpyxl
from dateutil.relativedelta import relativedelta
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


PIPEDRIVE_DOMAIN = "equationsie"

REQUIRED_FIELDS = {
    "id":      "ID (identifiant unique)",
    "societe": "Nom de la société",
    "date":    "Date d'entrée dans les lieux",
}
OPTIONAL_FIELDS = {
    "siren":   "SIREN",
    "adresse": "Adresse",
    "cp":      "Code postal",
    "ville":   "Ville",
}
DEFAULT_MAPPING = {
    "id": 1, "societe": 2, "siren": 3, "cp": 12, "ville": 13, "adresse": 28, "date": 30,
}

AUTO_COLS = [
    ("ID", 10), ("Lien Pipedrive", 16), ("Société", 26), ("Adresse", 38), ("CP", 8), ("Ville", 18),
    ("Entrée dans les lieux", 14), ("Fin de bail", 13), ("Mois restants", 10), ("SIREN", 12),
]
MANUAL_COLS = [
    ("Négo en charge", 18), ("Contact", 20), ("Téléphone", 14), ("Email", 26), ("Date appel", 12),
    ("Statut", 14), ("Relance le", 12), ("Notes", 40),
]
CAT_META = {
    "Fin < 6 mois": ("priorité HAUTE", "C00000"),
    "Fin 6-9 mois": ("priorité MOYENNE", "ED7D31"),
    "Fin 9-12 mois": ("à anticiper", "4472C4"),
}


def _parse_date(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        s = v.strip()
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                continue
    return None


def _months_diff(start, end):
    months = (end.year - start.year) * 12 + (end.month - start.month)
    if end.day < start.day:
        months -= 1
    return months


def read_headers(source_bytes: bytes) -> list[tuple[int, str]]:
    """Lit les headers (ligne 1) du fichier source."""
    wb = openpyxl.load_workbook(BytesIO(source_bytes), data_only=True, read_only=True)
    ws = wb["Données"] if "Données" in wb.sheetnames else wb.active
    headers = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v is not None:
            headers.append((c, str(v).strip()))
    wb.close()
    return headers


def _extract_rows(source_bytes: bytes, today, mapping):
    """Extrait les données en mode read_only puis ferme le workbook."""
    wb = openpyxl.load_workbook(BytesIO(source_bytes), data_only=True, read_only=True)
    ws = wb["Données"] if "Données" in wb.sheetnames else wb.active

    col_id   = mapping["id"]
    col_soc  = mapping["societe"]
    col_date = mapping["date"]
    col_siren = mapping.get("siren")
    col_addr  = mapping.get("adresse")
    col_cp    = mapping.get("cp")
    col_ville = mapping.get("ville")

    # Déterminer les colonnes max à lire
    all_cols = [col_id, col_soc, col_date]
    for c in [col_siren, col_addr, col_cp, col_ville]:
        if c:
            all_cols.append(c)
    max_col = max(all_cols)

    seen = set()
    rows = []
    row_num = 0
    for row_cells in ws.iter_rows(min_row=2, max_col=max_col, values_only=False):
        row_num += 1
        id_val = row_cells[col_id - 1].value if col_id - 1 < len(row_cells) else None
        if id_val is None or id_val in seen:
            continue
        # Ignorer les formules et erreurs
        if isinstance(id_val, str) and (id_val.startswith("=") or id_val.strip().startswith("#")):
            continue
        seen.add(id_val)

        date_raw = row_cells[col_date - 1].value if col_date - 1 < len(row_cells) else None
        if isinstance(date_raw, str) and (date_raw.startswith("=") or date_raw.strip().startswith("#")):
            date_raw = None
        d = _parse_date(date_raw)
        if d is None:
            continue

        months_since = _months_diff(d, today)
        n_cycles = months_since // 36 + 1
        end = d + relativedelta(months=n_cycles * 36)
        months_until = _months_diff(today, end)

        def _get(col):
            if col and col - 1 < len(row_cells):
                v = row_cells[col - 1].value
                if isinstance(v, str) and (v.startswith("=") or v.strip().startswith("#")):
                    return None
                return v
            return None

        rows.append({
            "id": id_val,
            "societe": _get(col_soc),
            "siren":   _get(col_siren),
            "adresse": _get(col_addr),
            "cp":      _get(col_cp),
            "ville":   _get(col_ville),
            "entree": d, "fin": end, "mois": months_until,
        })

    wb.close()  # Libère la mémoire du workbook source
    return rows


def _write_call_sheet(wb, sheet_name, data_list, today):
    tag, color = CAT_META[sheet_name]
    s = wb.create_sheet(sheet_name)
    total_cols = len(AUTO_COLS) + len(MANUAL_COLS)
    last_col = get_column_letter(total_cols)

    title_font = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
    header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    auto_fill = PatternFill("solid", start_color="1F4E78")
    manual_fill = PatternFill("solid", start_color="548235")
    border = Border(*(Side(style="thin", color="BFBFBF") for _ in range(4)))

    s.merge_cells(f"A1:{last_col}1")
    s["A1"] = f"📞 {sheet_name.upper()} — {tag} — {len(data_list)} sociétés"
    s["A1"].font = title_font
    s["A1"].fill = PatternFill("solid", start_color=color)
    s["A1"].alignment = Alignment(horizontal="center", vertical="center")
    s.row_dimensions[1].height = 30

    s.merge_cells(f"A2:{last_col}2")
    s["A2"] = (f"Généré le {today.strftime('%d/%m/%Y')}. "
               "Bleu = infos société. Vert = à remplir après appel.")
    s["A2"].font = Font(italic=True, size=10, color="595959")
    s["A2"].alignment = Alignment(horizontal="center", vertical="center")
    s.row_dimensions[2].height = 22

    header_row = 3
    for i, (label, width) in enumerate(AUTO_COLS + MANUAL_COLS, 1):
        c = s.cell(row=header_row, column=i, value=label)
        c.fill = auto_fill if i <= len(AUTO_COLS) else manual_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
        s.column_dimensions[get_column_letter(i)].width = width
    s.row_dimensions[header_row].height = 32

    manual_fill_cell = PatternFill("solid", start_color="EAF4E6")
    cell_font = Font(name="Calibri", size=10)
    link_font = Font(name="Calibri", size=10, color="0563C1", underline="single")
    for idx, r in enumerate(data_list):
        excel_row = header_row + 1 + idx
        values = [r["id"], None, r["societe"], r["adresse"], r["cp"], r["ville"],
                  r["entree"], r["fin"], r["mois"], r["siren"]]
        for i, v in enumerate(values, 1):
            c = s.cell(row=excel_row, column=i, value=v)
            c.border = border
            c.font = cell_font
            c.alignment = Alignment(vertical="center")
            if i == 2 and r["id"] is not None:
                c.value = "Ouvrir"
                c.hyperlink = f"https://{PIPEDRIVE_DOMAIN}.pipedrive.com/organization/{r['id']}"
                c.font = link_font
                c.alignment = Alignment(vertical="center", horizontal="center")
            elif i in (7, 8):
                c.number_format = "DD/MM/YYYY"
            elif i in (5, 10):
                c.number_format = "0;;;@"
        for j in range(1, len(MANUAL_COLS) + 1):
            col_i = len(AUTO_COLS) + j
            c = s.cell(row=excel_row, column=col_i)
            c.border = border
            c.fill = manual_fill_cell
            if MANUAL_COLS[j - 1][0] in ("Date appel", "Relance le"):
                c.number_format = "DD/MM/YYYY"

    extra_rows = 20
    last_data_row = header_row + len(data_list)
    for e in range(extra_rows):
        excel_row = last_data_row + 1 + e
        for i in range(1, total_cols + 1):
            c = s.cell(row=excel_row, column=i)
            c.border = border
            if i > len(AUTO_COLS):
                c.fill = manual_fill_cell

    mois_col = get_column_letter(9)
    mois_range = f"{mois_col}{header_row + 1}:{mois_col}{last_data_row + extra_rows}"
    for op, formula, bg, fg in [
        ("lessThan", ["3"], "F8CBAD", "9C0006"),
        ("between", ["3", "5"], "FFE699", "806000"),
        ("between", ["6", "8"], "FFF2CC", "7F6000"),
        ("between", ["9", "12"], "DDEBF7", "1F4E78"),
    ]:
        s.conditional_formatting.add(mois_range,
            CellIsRule(operator=op, formula=formula,
                       fill=PatternFill("solid", start_color=bg),
                       font=Font(bold=True, color=fg)))

    statut_letter = get_column_letter(len(AUTO_COLS) + 6)
    statut_range = f"{statut_letter}{header_row + 1}:{statut_letter}{last_data_row + extra_rows}"
    for val, bg, fg in [("OK signé", "C6EFCE", "006100"),
                         ("À rappeler", "FFEB9C", "9C5700"),
                         ("Refus", "FFC7CE", "9C0006"),
                         ("Injoignable", "D9D9D9", "595959")]:
        s.conditional_formatting.add(statut_range,
            CellIsRule(operator="equal", formula=[f'"{val}"'],
                       fill=PatternFill("solid", start_color=bg),
                       font=Font(bold=True, color=fg)))
    dv = DataValidation(type="list",
        formula1='"À appeler,À rappeler,OK signé,Refus,Injoignable,Ne pas contacter"',
        allow_blank=True)
    dv.add(statut_range)
    s.add_data_validation(dv)

    s.auto_filter.ref = f"A{header_row}:{last_col}{last_data_row + extra_rows}"
    s.freeze_panes = s.cell(row=header_row + 1, column=4)


def _write_readme(wb, today, counts):
    r = wb.create_sheet("Mode d'emploi", 0)
    r.column_dimensions["A"].width = 28
    r.column_dimensions["B"].width = 100
    r["A1"] = "📞 Suivi commercial — fins de bail triennales"
    r.merge_cells("A1:B1")
    r["A1"].font = Font(bold=True, size=16, color="FFFFFF", name="Calibri")
    r["A1"].fill = PatternFill("solid", start_color="1F4E78")
    r["A1"].alignment = Alignment(horizontal="center", vertical="center")
    r.row_dimensions[1].height = 34

    total = sum(counts.values())
    lines = [
        ("", ""),
        ("Date de calcul", f"{today.strftime('%d/%m/%Y')} — {total} sociétés au total."),
        ("", ""),
        ("🔴 Fin < 6 mois", f"{counts['Fin < 6 mois']} sociétés — priorité HAUTE."),
        ("🟠 Fin 6-9 mois", f"{counts['Fin 6-9 mois']} sociétés — priorité MOYENNE."),
        ("🔵 Fin 9-12 mois", f"{counts['Fin 9-12 mois']} sociétés — à anticiper."),
        ("", ""),
        ("Colonnes bleues", "Auto (ID, Lien Pipedrive, Société, Adresse, CP, Ville, Entrée, Fin de bail, Mois restants, SIREN)."),
        ("Colonnes vertes", "À remplir (Négo en charge, Contact, Téléphone, Email, Date appel, Statut, Relance, Notes)."),
        ("Statut", "Liste déroulante : À appeler, À rappeler, OK signé, Refus, Injoignable, Ne pas contacter."),
        ("Code couleur Mois", "Rouge < 3 • Orange 3-5 • Jaune 6-8 • Bleu 9-12."),
    ]
    for i, (k, v) in enumerate(lines, 2):
        r.cell(row=i, column=1, value=k).font = Font(bold=True, name="Calibri", size=11)
        r.cell(row=i, column=1).alignment = Alignment(vertical="top", wrap_text=True)
        r.cell(row=i, column=2, value=v).alignment = Alignment(wrap_text=True, vertical="top")
        r.cell(row=i, column=2).font = Font(name="Calibri", size=11)
        r.row_dimensions[i].height = 26


def generate(source_bytes: bytes, mapping: dict = None, today=None) -> tuple[bytes, dict]:
    """Extrait les données (read_only), puis construit un nouveau workbook pour l'output."""
    today = today or date.today()
    mapping = mapping or DEFAULT_MAPPING

    # Phase 1 : extraction (read_only → libère la RAM du source)
    rows = _extract_rows(source_bytes, today, mapping)

    # Phase 2 : catégorisation
    cat_rows = {
        "Fin < 6 mois":  sorted([r for r in rows if r["mois"] < 6], key=lambda x: (x["mois"], x["fin"])),
        "Fin 6-9 mois":  sorted([r for r in rows if 6 <= r["mois"] < 9], key=lambda x: (x["mois"], x["fin"])),
        "Fin 9-12 mois": sorted([r for r in rows if 9 <= r["mois"] <= 12], key=lambda x: (x["mois"], x["fin"])),
    }
    del rows  # Libère la liste complète
    counts = {k: len(v) for k, v in cat_rows.items()}

    # Phase 3 : nouveau workbook (pas de rechargement du source)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Supprime la feuille par défaut

    for sheet_name, data_list in cat_rows.items():
        _write_call_sheet(wb, sheet_name, data_list, today)
    _write_readme(wb, today, counts)

    out = BytesIO()
    wb.save(out)
    wb.close()
    return out.getvalue(), cat_rows
