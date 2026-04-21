"""
auto_refresh.py — Process mensuel automatique.

Lit les fichiers *générés* depuis Dropbox (3 onglets : Fin < 6 mois, 6-9, 9-12),
extrait les données brutes (ID, société, adresse, dates, SIREN...),
recalcule les mois restants à la date du jour,
re-catégorise, génère un nouveau fichier, sync Notion, envoie par mail.
"""

import gc
import logging
from datetime import date, datetime
from io import BytesIO

import openpyxl
from dateutil.relativedelta import relativedelta

log = logging.getLogger("auto_refresh")

PIPEDRIVE_DOMAIN = "equationsie"

SHEETS_TO_READ = ["Fin < 6 mois", "Fin 6-9 mois", "Fin 9-12 mois"]

# Colonnes des fichiers générés (toujours les mêmes) :
# A=ID, B=Lien Pipedrive, C=Société, D=Adresse, E=CP, F=Ville,
# G=Entrée dans les lieux, H=Fin de bail, I=Mois restants, J=SIREN
# Headers en ligne 3, data à partir de ligne 4


def _parse_date(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        s = v.strip()
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                continue
    return None


def _months_diff(start, end):
    months = (end.year - start.year) * 12 + (end.month - start.month)
    if end.day < start.day:
        months -= 1
    return months


def extract_from_generated(files: list[tuple[str, bytes]], today: date) -> list[dict]:
    """
    Lit les fichiers générés, extrait les données brutes,
    recalcule les mois restants à partir de la date d'entrée.
    Dédoublonne par ID (garde la tranche la plus urgente).
    """
    seen = {}  # id -> row dict

    for filename, content in files:
        log.info(f"  Lecture de {filename}...")
        try:
            wb = openpyxl.load_workbook(BytesIO(content), data_only=True, read_only=True)
        except Exception as e:
            log.error(f"  ✗ Impossible d'ouvrir {filename}: {e}")
            continue

        for sheet_name in SHEETS_TO_READ:
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]

            for row_cells in ws.iter_rows(min_row=4, max_col=10, values_only=False):
                if len(row_cells) < 10:
                    continue

                id_val = row_cells[0].value  # col A = ID
                if id_val is None:
                    continue
                id_str = str(id_val).strip()
                if not id_str:
                    continue

                # Date d'entrée (col G = index 6)
                entree = _parse_date(row_cells[6].value)
                if entree is None:
                    continue

                # Recalculer la fin de bail à partir de la date d'entrée
                months_since = _months_diff(entree, today)
                n_cycles = months_since // 36 + 1
                fin = entree + relativedelta(months=n_cycles * 36)
                mois = _months_diff(today, fin)

                row = {
                    "id": id_val,
                    "societe": row_cells[2].value,   # col C
                    "adresse": row_cells[3].value,   # col D
                    "cp":      row_cells[4].value,   # col E
                    "ville":   row_cells[5].value,   # col F
                    "siren":   row_cells[9].value,   # col J
                    "entree":  entree,
                    "fin":     fin,
                    "mois":    mois,
                }

                # Dédoublonnage : garder la version avec le moins de mois restants
                prev = seen.get(id_str)
                if prev is None or mois < prev["mois"]:
                    seen[id_str] = row

        wb.close()

    rows = list(seen.values())
    log.info(f"  {len(rows)} sociétés uniques extraites (recalculées au {today.strftime('%d/%m/%Y')})")
    return rows


def categorize(rows: list[dict]) -> dict:
    """Répartit les rows dans les 3 catégories."""
    return {
        "Fin < 6 mois":  sorted([r for r in rows if r["mois"] < 6],
                                 key=lambda x: (x["mois"], x["fin"])),
        "Fin 6-9 mois":  sorted([r for r in rows if 6 <= r["mois"] < 9],
                                 key=lambda x: (x["mois"], x["fin"])),
        "Fin 9-12 mois": sorted([r for r in rows if 9 <= r["mois"] <= 12],
                                 key=lambda x: (x["mois"], x["fin"])),
    }
